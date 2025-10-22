"""
Manejador de archivos Excel para guardar y recuperar calificaciones
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os


class ExcelHandler:
    """
    Maneja la lectura y escritura de calificaciones en archivos Excel
    """
    
    def __init__(self, filepath):
        """
        Inicializa el manejador de Excel
        
        Args:
            filepath: Ruta del archivo Excel
        """
        self.filepath = filepath
        self.workbook = None
        self.sheet = None
        self.students = {}  # {matricula: nombre}
        self.matricula_col = 1  # Columna A
        self.name_col = 2  # Columna B
        
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"El archivo {filepath} no existe")
        
        self.load_workbook()
    
    def load_workbook(self):
        """Carga el archivo Excel y los datos de estudiantes"""
        try:
            self.workbook = openpyxl.load_workbook(self.filepath)
            self.sheet = self.workbook.active
            
            # Cargar estudiantes (desde fila 2 en adelante)
            for row in range(2, self.sheet.max_row + 1):
                matricula = self.sheet.cell(row, self.matricula_col).value
                nombre = self.sheet.cell(row, self.name_col).value
                
                if matricula and nombre:
                    self.students[str(matricula)] = {
                        'name': nombre,
                        'row': row
                    }
            
            print(f"✓ Excel cargado: {len(self.students)} estudiantes encontrados")
            
        except Exception as e:
            raise Exception(f"Error al cargar el archivo Excel: {e}")
    
    def get_student_by_matricula(self, matricula):
        """
        Busca un estudiante por su matrícula
        
        Args:
            matricula: Número de matrícula del estudiante
        
        Returns:
            dict o None: Información del estudiante si existe
        """
        return self.students.get(str(matricula))
    
    def find_or_create_test_column(self, test_name):
        """
        Encuentra o crea una columna para la prueba especificada
        
        Args:
            test_name: Nombre de la prueba
        
        Returns:
            int: Número de columna
        """
        # Buscar si ya existe una columna con este nombre
        for col in range(3, self.sheet.max_column + 2):
            cell_value = self.sheet.cell(1, col).value
            if cell_value == test_name:
                return col
        
        # Si no existe, crear nueva columna
        new_col = self.sheet.max_column + 1
        
        # Agregar encabezado con formato
        header_cell = self.sheet.cell(1, new_col)
        header_cell.value = test_name
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", 
                                       fill_type="solid")
        
        return new_col
    
    def check_existing_grade(self, matricula, test_name):
        """
        Verifica si un estudiante ya tiene nota para una prueba
        
        Args:
            matricula: Número de matrícula
            test_name: Nombre de la prueba
        
        Returns:
            tuple: (exists: bool, grade: float o None)
        """
        student = self.get_student_by_matricula(matricula)
        if not student:
            return False, None
        
        col = self.find_or_create_test_column(test_name)
        row = student['row']
        
        existing_grade = self.sheet.cell(row, col).value
        
        if existing_grade is not None and existing_grade != "":
            return True, existing_grade
        
        return False, None
    
    def save_grade(self, matricula, test_name, grade, overwrite=False):
        """
        Guarda la nota de un estudiante
        
        Args:
            matricula: Número de matrícula
            test_name: Nombre de la prueba
            grade: Nota a guardar
            overwrite: Si es True, sobrescribe nota existente
        
        Returns:
            dict: Resultado de la operación
        """
        # Verificar que el estudiante existe
        student = self.get_student_by_matricula(matricula)
        if not student:
            return {
                'success': False,
                'message': f"Estudiante con matrícula {matricula} no encontrado"
            }
        
        # Verificar si ya existe una nota
        exists, existing_grade = self.check_existing_grade(matricula, test_name)
        
        if exists and not overwrite:
            return {
                'success': False,
                'message': 'Ya existe una nota para este estudiante',
                'existing_grade': existing_grade,
                'requires_confirmation': True
            }
        
        # Guardar la nota
        try:
            col = self.find_or_create_test_column(test_name)
            row = student['row']
            
            cell = self.sheet.cell(row, col)
            cell.value = grade
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Aplicar color según aprobación (asumiendo 4.0 como nota de aprobación)
            if grade >= 4.0:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", 
                                       fill_type="solid")  # Verde claro
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", 
                                       fill_type="solid")  # Rojo claro
            
            # Guardar el archivo
            self.workbook.save(self.filepath)
            
            return {
                'success': True,
                'message': 'Nota guardada exitosamente',
                'student_name': student['name'],
                'grade': grade,
                'was_overwritten': exists
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f"Error al guardar la nota: {e}"
            }
    
    def get_all_students(self):
        """
        Obtiene la lista de todos los estudiantes
        
        Returns:
            dict: Diccionario con todos los estudiantes
        """
        return self.students
    
    def close(self):
        """Cierra el archivo Excel"""
        if self.workbook:
            self.workbook.close()