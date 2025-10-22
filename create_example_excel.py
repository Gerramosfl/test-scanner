"""
Script para crear un archivo Excel de ejemplo con lista de estudiantes
Ejecutar: python create_example_excel.py
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

def create_example_excel():
    """Crea un archivo Excel de ejemplo con estudiantes"""
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estudiantes"
    
    # Configurar encabezados
    headers = ["Matrícula", "Nombre Alumno"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", 
                               fill_type="solid")
    
    # Datos de estudiantes de ejemplo (basados en tu Excel)
    students = [
        ("2023456195", "Aedo Acuña Antonia Valentina"),
        ("2023418927", "Alarcón Llanos Daniela Annaís"),
        ("2022044110", "Alarcón Mendoza Sofía Valentina"),
        ("2022436191", "Aqueveque Fuentes Javier Alexis"),
        ("2023412139", "Aravena Pardo Isis Alexandra"),
        ("2023430129", "Avilez Placencia Martina Beatriz"),
        ("2023440949", "Badilla Torres Isidora Belén Minerva"),
        ("2023455822", "Bianchi Quevedo Mariyna Ignacia"),
        ("2023442194", "Bravo Ramírez Florencia Isabel"),
        ("2023413470", "Campos Veloso Antonella Soleil"),
        ("2023437611", "Carrasco Regla Martín Antonio"),
        ("2023432253", "Castillo Leal Jaime Alejandro"),
        ("2022430281", "Castillo Oliva Francisco Javier"),
        ("2023408611", "Cerpa Navarrete Violeta Almendra"),
        ("2023406970", "Césped Leal Fernando Andrés"),
        ("2023406860", "Chamorro Sanhueza Carlos Vicente"),
        ("2023424315", "Contreras Rubilar Felipe Andrés"),
        ("2023444592", "Cuevas Gutiérrez Catalina Anaís"),
        ("2022750375", "Fierro Guzmán Paloma Josefa"),
        ("2023416029", "Figueroa Vásquez Bastian Nicolás"),
    ]
    
    # Agregar estudiantes
    for row, (matricula, nombre) in enumerate(students, start=2):
        ws.cell(row, 1).value = matricula
        ws.cell(row, 2).value = nombre
        
        # Alinear contenido
        ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row, 2).alignment = Alignment(horizontal='left', vertical='center')
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    
    # Crear directorio examples si no existe
    os.makedirs('examples', exist_ok=True)
    
    # Guardar archivo
    filepath = 'examples/lista_alumnos_ejemplo.xlsx'
    wb.save(filepath)
    
    print(f"✓ Archivo creado exitosamente: {filepath}")
    print(f"  Total de estudiantes: {len(students)}")
    print("\nPuedes usar este archivo como plantilla para tus propios cursos.")
    print("Solo necesitas tener las columnas 'Matrícula' y 'Nombre Alumno'.")

if __name__ == "__main__":
    create_example_excel()